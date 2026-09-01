"""Экспорт паккинг-листа в формате ATA Carnet General List (Excel).

Заявляемая стоимость (Value) в системе не хранится — это не арендная цена, а
таможенная/страховая стоимость товара, которую нужно вписать вручную после
экспорта. Колонка «Стоимость» в файле оставляется пустой.

Серийный номер (S/N) в списке — штрих-код назначенного экземпляра. Если у
строки нет отслеживаемых экземпляров (количественное оборудование), пишем
«NSN» (No Serial Number), как того требует формат ATA Carnet.
"""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.accessory_kits import service as accessory_kit_service
from app.accessory_kits.models import AccessoryKitLine
from app.inventory.models import EquipmentModel
from app.inventory.services import kits as kit_service
from app.inventory.services.kits import KitGroup
from app.packing.models import PackingLine, PackingList
from app.projects.models import Project


@dataclass
class CarnetRow:
    description: str
    quantity: int
    weight_kg: Decimal
    country_of_origin: str


def _models_by_id(db: Session, packing: PackingList) -> dict[int, EquipmentModel]:
    ids = {ln.model_id for ln in packing.lines if ln.model_id is not None}
    if not ids:
        return {}
    stmt = select(EquipmentModel).where(EquipmentModel.id.in_(ids))
    return {m.id: m for m in db.execute(stmt).scalars().all()}


def _accessory_content_models(db: Session, packing: PackingList) -> dict[int, EquipmentModel]:
    """Модели содержимого комплектов аксессуаров проекта (для производителя/страны в carnet)."""
    ids: set[int] = set()
    for ln in packing.lines:
        if ln.accessory_kit_id is None:
            continue
        kit = accessory_kit_service.get(db, ln.accessory_kit_id)
        if kit is None:
            continue
        ids.update(cl.model_id for cl in kit.lines if cl.model_id is not None)
    if not ids:
        return {}
    stmt = select(EquipmentModel).where(EquipmentModel.id.in_(ids))
    return {m.id: m for m in db.execute(stmt).scalars().all()}


def _serial_text(barcodes: list[str]) -> str:
    barcodes = [b for b in barcodes if b]
    return "S/N: " + ", ".join(barcodes) if barcodes else "NSN"


def _carnet_barcodes(line: PackingLine) -> list[str]:
    """Штрих-коды строки для carnet: подтверждённые сканом, если такие есть —
    иначе все имеющиеся (включая незаменённые сканом заготовки, ТЗ §22)."""
    confirmed = [si.barcode for si in line.serial_items if si.confirmed_by_scan]
    return confirmed if confirmed else [si.barcode for si in line.serial_items]


def _model_line_row(line: PackingLine, model: EquipmentModel | None, qty: int) -> CarnetRow:
    # Количественное оборудование, добавленное/дособранное сканом, тоже несёт
    # штрих-коды в packing_serial_items — используем их и здесь, не только у
    # серийных строк (line.is_serial=False не значит «без штрих-кодов»).
    serial_text = _serial_text(_carnet_barcodes(line))
    desc = line.name
    if model and model.manufacturer:
        desc += f", {model.manufacturer}"
    desc += f", {serial_text}"
    country = (model.country_of_origin if model else None) or ""
    return CarnetRow(desc, qty, line.unit_weight_kg * qty, country)


def _accessory_content_row(cl: AccessoryKitLine, model: EquipmentModel | None) -> CarnetRow:
    """Строка carnet для позиции содержимого комплекта аксессуаров.

    Содержимое количественное (модель+кол-во, без отслеживаемых экземпляров) —
    в отличие от Комплекта здесь нет штрих-кодов отдельных единиц, поэтому S/N
    всегда «NSN», как того требует формат ATA Carnet для неотслеживаемых позиций.
    """
    desc = cl.name
    if model and model.manufacturer:
        desc += f", {model.manufacturer}"
    desc += ", NSN"
    country = (model.country_of_origin if model else None) or ""
    return CarnetRow(desc, cl.quantity, cl.unit_weight_kg * cl.quantity, country)


def _kit_group_row(group: KitGroup) -> CarnetRow:
    model = group.items[0].model if group.items else None
    serial_text = _serial_text([it.barcode for it in group.items])
    desc = group.model_name
    if model and model.manufacturer:
        desc += f", {model.manufacturer}"
    desc += f", {serial_text}"
    country = (model.country_of_origin if model else None) or ""
    weight = (model.weight_kg if model else Decimal("0")) * group.count
    return CarnetRow(desc, group.count, weight, country)


def build_rows(db: Session, packing: PackingList) -> list[CarnetRow]:
    """Построчный список для carnet: комплекты и комплекты аксессуаров разворачиваются
    по составу (ATA требует поштучного/по-позиционного описания, группировка допустима
    только для одинаковых единиц — ТЗ на комплекты этому уже соответствует, см. `KitGroup`).
    Комплект аксессуаров в смете виден одной строкой без содержимого (ТЗ), но в carnet
    декларируется реальное количество провозимого — поэтому здесь именно содержимое."""
    models = _models_by_id(db, packing)
    accessory_models = _accessory_content_models(db, packing)
    rows: list[CarnetRow] = []
    ordered = sorted(packing.lines, key=lambda ln: (ln.sort_order, ln.id))
    for line in ordered:
        if line.kit_id is not None:
            kit = kit_service.get_kit(db, line.kit_id)
            if kit is not None:
                for group in kit_service.content_groups(kit):
                    if group.count > 0:
                        rows.append(_kit_group_row(group))
            continue
        if line.accessory_kit_id is not None:
            ak = accessory_kit_service.get(db, line.accessory_kit_id)
            if ak is not None:
                for cl in ak.lines:
                    if cl.quantity > 0:
                        model = accessory_models.get(cl.model_id) if cl.model_id else None
                        rows.append(_accessory_content_row(cl, model))
            continue
        qty = line.fact_quantity
        if qty <= 0:
            continue
        model = models.get(line.model_id) if line.model_id else None
        rows.append(_model_line_row(line, model, qty))
    return rows


def build_workbook(project: Project, packing: PackingList, rows: list[CarnetRow]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "General List"

    bold = Font(bold=True)
    ws["A1"] = "ATA CARNET — GENERAL LIST"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"{project.number} · {project.name}"
    ws["A3"] = f"Packing-лист: {packing.number}"
    ws["A4"] = "Стоимость (Value) заполняется вручную после экспорта."
    ws["A4"].font = Font(italic=True, color="808080")

    headers = [
        "№ / Item No.",
        "Кол-во мест / Pieces",
        "Описание (марка, S/N) / Description",
        "Страна происх. / Origin",
        "Вес, кг / Weight, kg",
        "Стоимость / Value",
    ]
    header_row = 6
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font = bold

    r = header_row + 1
    for i, row in enumerate(rows, start=1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=row.quantity)
        ws.cell(row=r, column=3, value=row.description)
        ws.cell(row=r, column=4, value=row.country_of_origin)
        ws.cell(row=r, column=5, value=float(row.weight_kg))
        ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    if rows:
        total_row = r + 1
        ws.cell(row=total_row, column=2, value="Итого / Total:").font = bold
        ws.cell(row=total_row, column=5, value=f"=SUM(E{header_row + 1}:E{r - 1})")

    widths = [6, 14, 60, 16, 14, 16]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    return wb
