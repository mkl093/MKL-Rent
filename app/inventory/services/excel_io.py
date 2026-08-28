"""Экспорт карточки модели в Excel и обратный импорт (параметры модели + единицы).

Файл — две страницы: «Карточка товара» (ключ/значение + комплектация) и
«Единицы оборудования» (таблица экземпляров). Фото, мануалы и сертификаты —
файлы, не табличные данные, поэтому не экспортируются.

Импорт: только обновление уже существующей модели (ID в файле должен совпадать
с открытой карточкой), тип учёта неизменяем. Валидация — в два прохода: сперва
читаем и проверяем весь файл целиком, и лишь если ошибок нет — применяем
изменения. Строки листа «Единицы оборудования», отсутствующие в файле,
не трогаем (единицы не удаляются импортом).
"""

from __future__ import annotations

import io
import zipfile
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.database import utcnow
from app.inventory.enums import AccountingType, ItemStatus, PackingType
from app.inventory.models import (
    Category,
    EquipmentItem,
    EquipmentModel,
    EquipmentStatusHistory,
    Subcategory,
)
from app.inventory.schemas import AccessoryQty, EquipmentModelUpdate, PackingRuleInput
from app.inventory.services import accessories as acc_service
from app.inventory.services import categories as cat_service
from app.inventory.services import equipment as eq_service
from app.inventory.services import items as item_service
from app.inventory.services.categories import InventoryError

CARD_SHEET = "Карточка товара"
ITEMS_SHEET = "Единицы оборудования"

FIELD_LABELS: dict[str, str] = {
    "id": "ID модели",
    "category": "Категория",
    "subcategory": "Подкатегория",
    "name": "Название",
    "accounting_type": "Тип учёта",
    "manufacturer": "Производитель",
    "country_of_origin": "Страна производства",
    "internal_sku": "Артикул",
    "storage_location": "Место хранения",
    "weight_kg": "Вес, кг",
    "length_mm": "Длина, мм",
    "width_mm": "Ширина, мм",
    "height_mm": "Высота, мм",
    "base_price_eur": "Базовая цена, €",
    "total_quantity": "Количество (только для количественного типа)",
    "has_power": "Есть энергопотребление",
    "power_peak_w": "Пиковая мощность, Вт",
    "power_nominal_w": "Номинальная мощность, Вт",
    "packing_type": "Упаковка: тип (Кейс/Рэк)",
    "packing_empty_weight_kg": "Упаковка: вес тары, кг",
    "packing_length_mm": "Упаковка: длина, мм",
    "packing_width_mm": "Упаковка: ширина, мм",
    "packing_height_mm": "Упаковка: высота, мм",
    "packing_capacity": "Упаковка: вместимость, шт",
    "description": "Описание",
    "note": "Примечание",
}
FIELD_ORDER = list(FIELD_LABELS.keys())

ACCESSORY_HEADER = ["Категория аксессуара", "Аксессуар", "Количество"]
ITEM_HEADERS = [
    "Модель",
    "ID",
    "Штрих-код",
    "Серийный номер",
    "Инвентарный номер",
    "Статус",
    "Несмотря на дефект",
    "Комментарий",
    "Комплект",
]

STATUS_BY_LABEL: dict[str, ItemStatus] = {s.label.lower(): s for s in ItemStatus} | {
    s.value: s for s in ItemStatus
}
PACKING_BY_LABEL: dict[str, PackingType] = {p.label.lower(): p for p in PackingType} | {
    p.value: p for p in PackingType
}


class ImportValidationError(InventoryError):
    """Файл не прошёл проверку — импорт не применён (список ошибок в .errors)."""

    def __init__(self, errors: list[str]):
        super().__init__("; ".join(errors))
        self.errors = errors


@dataclass
class ImportResult:
    items_created: int
    items_updated: int
    quantity_adjusted: bool


# --- Экспорт --------------------------------------------------------------


def build_workbook(db: Session, model: EquipmentModel) -> Workbook:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = CARD_SHEET
    _write_card_sheet(ws1, model)
    ws2 = wb.create_sheet(ITEMS_SHEET)
    _write_items_sheet(ws2, db, model)
    return wb


def _write_card_sheet(ws, model: EquipmentModel) -> None:
    bold = Font(bold=True)
    values: dict[str, object] = {
        "id": model.id,
        "category": model.category.name if model.category else "",
        "subcategory": model.subcategory.name if model.subcategory else "",
        "name": model.name,
        "accounting_type": model.accounting_type.label,
        "manufacturer": model.manufacturer or "",
        "country_of_origin": model.country_of_origin or "",
        "internal_sku": model.internal_sku or "",
        "storage_location": model.storage_location or "",
        "weight_kg": float(model.weight_kg),
        "length_mm": model.length_mm,
        "width_mm": model.width_mm,
        "height_mm": model.height_mm,
        "base_price_eur": float(model.base_price_eur),
        "total_quantity": model.total_quantity
        if model.accounting_type == AccountingType.QUANTITY
        else "",
        "has_power": "Да" if model.has_power else "Нет",
        "power_peak_w": model.power_peak_w if model.has_power else "",
        "power_nominal_w": model.power_nominal_w if model.has_power else "",
        "packing_type": model.packing.packing_type.label if model.packing else "",
        "packing_empty_weight_kg": float(model.packing.empty_weight_kg) if model.packing else "",
        "packing_length_mm": model.packing.length_mm if model.packing else "",
        "packing_width_mm": model.packing.width_mm if model.packing else "",
        "packing_height_mm": model.packing.height_mm if model.packing else "",
        "packing_capacity": model.packing.capacity if model.packing else "",
        "description": model.description or "",
        "note": model.note or "",
    }
    row = 1
    for key in FIELD_ORDER:
        ws.cell(row=row, column=1, value=FIELD_LABELS[key]).font = bold
        ws.cell(row=row, column=2, value=values[key])
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Комплектация (аксессуары)").font = bold
    row += 1
    for col, label in enumerate(ACCESSORY_HEADER, start=1):
        ws.cell(row=row, column=col, value=label).font = bold
    row += 1
    for ma in model.accessories:
        ws.cell(row=row, column=1, value=ma.accessory.category.name)
        ws.cell(row=row, column=2, value=ma.accessory.name)
        ws.cell(row=row, column=3, value=ma.quantity)
        row += 1

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14


def _write_items_sheet(ws, db: Session, model: EquipmentModel) -> None:
    bold = Font(bold=True)
    for col, label in enumerate(ITEM_HEADERS, start=1):
        ws.cell(row=1, column=col, value=label).font = bold
    items = item_service.list_items(db, model.id)
    for i, it in enumerate(items, start=2):
        ws.cell(row=i, column=1, value=model.name)
        ws.cell(row=i, column=2, value=it.id)
        ws.cell(row=i, column=3, value=it.barcode or "")
        ws.cell(row=i, column=4, value=it.serial_number or "")
        ws.cell(row=i, column=5, value=it.inventory_number or "")
        ws.cell(row=i, column=6, value=it.status.label)
        ws.cell(row=i, column=7, value="Да" if it.usable_despite_defect else "Нет")
        ws.cell(row=i, column=8, value=it.comment or "")
        ws.cell(row=i, column=9, value=it.kit.name if it.kit else "")
    widths = [28, 8, 22, 20, 20, 14, 18, 32, 24]
    for col_letter, width in zip("ABCDEFGHI", widths, strict=True):
        ws.column_dimensions[col_letter].width = width


# --- Импорт: разбор ячеек --------------------------------------------------


def _cell_str(ws, row: int, col: int) -> str:
    value = ws.cell(row=row, column=col).value
    if value is None:
        return ""
    return str(value).strip()


def _parse_bool(raw: str) -> bool:
    return raw.strip().lower() in ("да", "yes", "true", "1", "истина")


def _parse_int_field(
    raw: str, label: str, errors: list[str], *, minimum: int, default: int = 0
) -> int:
    raw = raw.strip()
    if not raw:
        return default
    try:
        value = int(float(raw.replace(",", ".")))
    except ValueError:
        errors.append(f"Некорректное число в поле «{label}»: «{raw}».")
        return default
    if value < minimum:
        errors.append(f"«{label}» не может быть меньше {minimum}.")
        return default
    return value


def _parse_decimal_field(
    raw: str,
    label: str,
    errors: list[str],
    *,
    minimum: Decimal = Decimal("0"),
    default: Decimal = Decimal("0"),
) -> Decimal:
    raw = raw.strip()
    if not raw:
        return default
    try:
        value = Decimal(raw.replace(",", "."))
    except InvalidOperation:
        errors.append(f"Некорректное число в поле «{label}»: «{raw}».")
        return default
    if value < minimum:
        errors.append(f"«{label}» не может быть меньше {minimum}.")
        return default
    return value


@dataclass
class _ItemPlan:
    row: int
    existing: EquipmentItem | None
    barcode: str | None
    serial_number: str | None
    inventory_number: str | None
    comment: str | None
    status: ItemStatus
    usable_despite_defect: bool


def _load_workbook(file_bytes: bytes) -> Workbook:
    try:
        return load_workbook(io.BytesIO(file_bytes), data_only=True)
    except (InvalidFileException, OSError, KeyError, zipfile.BadZipFile) as exc:
        raise ImportValidationError(
            ["Не удалось прочитать файл — это не похоже на файл .xlsx или он повреждён."]
        ) from exc


def _read_card_fields(
    ws, db: Session, model: EquipmentModel, errors: list[str]
) -> tuple[EquipmentModelUpdate | None, int]:
    """Разобрать лист «Карточка товара». Возвращает (данные модели, количество)."""
    label_to_row: dict[str, int] = {}
    for r in range(1, ws.max_row + 1):
        label = _cell_str(ws, r, 1)
        if label and label not in label_to_row:
            label_to_row[label] = r

    def cell_value(key: str) -> str:
        r = label_to_row.get(FIELD_LABELS[key])
        if r is None:
            return ""
        return _cell_str(ws, r, 2)

    id_raw = cell_value("id")
    try:
        file_model_id = int(float(id_raw)) if id_raw else None
    except ValueError:
        file_model_id = None
    if file_model_id != model.id:
        errors.append(
            f"ID модели в файле ({id_raw or '—'}) не совпадает с открытой карточкой "
            f"(ID {model.id}). Импорт поддерживает только обновление той же модели."
        )

    name = cell_value("name")
    if not name:
        errors.append(f"Не заполнено «{FIELD_LABELS['name']}».")

    accounting_label = cell_value("accounting_type")
    if accounting_label and accounting_label.lower() != model.accounting_type.label.lower():
        errors.append(
            "Тип учёта нельзя изменить импортом "
            f"(в файле «{accounting_label}», у модели «{model.accounting_type.label}»)."
        )

    category_name = cell_value("category")
    category: Category | None = None
    if not category_name:
        errors.append(f"Не заполнена «{FIELD_LABELS['category']}».")
    else:
        category = _find_by_name(cat_service.list_categories(db), category_name)
        if category is None:
            errors.append(f"Категория «{category_name}» не найдена в справочнике.")

    subcategory_name = cell_value("subcategory")
    subcategory: Subcategory | None = None
    if subcategory_name and category is not None:
        subcategory = _find_by_name(category.subcategories, subcategory_name)
        if subcategory is None:
            errors.append(
                f"Подкатегория «{subcategory_name}» не найдена в категории «{category_name}»."
            )

    weight_kg = _parse_decimal_field(cell_value("weight_kg"), FIELD_LABELS["weight_kg"], errors)
    length_mm = _parse_int_field(
        cell_value("length_mm"), FIELD_LABELS["length_mm"], errors, minimum=0
    )
    width_mm = _parse_int_field(
        cell_value("width_mm"), FIELD_LABELS["width_mm"], errors, minimum=0
    )
    height_mm = _parse_int_field(
        cell_value("height_mm"), FIELD_LABELS["height_mm"], errors, minimum=0
    )
    base_price_eur = _parse_decimal_field(
        cell_value("base_price_eur"), FIELD_LABELS["base_price_eur"], errors
    )

    total_quantity = 0
    if model.accounting_type == AccountingType.QUANTITY:
        total_quantity = _parse_int_field(
            cell_value("total_quantity"), FIELD_LABELS["total_quantity"], errors, minimum=0
        )

    has_power = _parse_bool(cell_value("has_power"))
    power_peak_w: int | None = None
    power_nominal_w: int | None = None
    if has_power:
        peak_raw = cell_value("power_peak_w")
        nominal_raw = cell_value("power_nominal_w")
        if not peak_raw or not nominal_raw:
            errors.append(
                "При «Есть энергопотребление: Да» нужно заполнить пиковую и номинальную мощность."
            )
        else:
            power_peak_w = _parse_int_field(
                peak_raw, FIELD_LABELS["power_peak_w"], errors, minimum=0
            )
            power_nominal_w = _parse_int_field(
                nominal_raw, FIELD_LABELS["power_nominal_w"], errors, minimum=0
            )

    packing_type_label = cell_value("packing_type")
    packing_input: PackingRuleInput | None = None
    if packing_type_label:
        ptype = PACKING_BY_LABEL.get(packing_type_label.strip().lower())
        if ptype is None:
            errors.append(
                f"Неизвестный тип упаковки «{packing_type_label}» (ожидается «Кейс» или «Рэк»)."
            )
        else:
            empty_weight = _parse_decimal_field(
                cell_value("packing_empty_weight_kg"),
                FIELD_LABELS["packing_empty_weight_kg"],
                errors,
            )
            p_length = _parse_int_field(
                cell_value("packing_length_mm"),
                FIELD_LABELS["packing_length_mm"],
                errors,
                minimum=0,
            )
            p_width = _parse_int_field(
                cell_value("packing_width_mm"),
                FIELD_LABELS["packing_width_mm"],
                errors,
                minimum=0,
            )
            p_height = _parse_int_field(
                cell_value("packing_height_mm"),
                FIELD_LABELS["packing_height_mm"],
                errors,
                minimum=0,
            )
            capacity = _parse_int_field(
                cell_value("packing_capacity"),
                FIELD_LABELS["packing_capacity"],
                errors,
                minimum=1,
                default=1,
            )
            packing_input = PackingRuleInput(
                packing_type=ptype,
                empty_weight_kg=empty_weight,
                length_mm=p_length,
                width_mm=p_width,
                height_mm=p_height,
                capacity=capacity,
            )

    accessory_items = _read_accessories(ws, db, errors)

    if errors:
        return None, total_quantity

    data = EquipmentModelUpdate(
        category_id=category.id,  # type: ignore[union-attr]
        name=name,
        weight_kg=weight_kg,
        length_mm=length_mm,
        width_mm=width_mm,
        height_mm=height_mm,
        base_price_eur=base_price_eur,
        total_quantity=model.total_quantity,
        subcategory_id=subcategory.id if subcategory else None,
        manufacturer=cell_value("manufacturer") or None,
        country_of_origin=cell_value("country_of_origin") or None,
        internal_sku=cell_value("internal_sku") or None,
        description=cell_value("description") or None,
        note=cell_value("note") or None,
        storage_location=cell_value("storage_location") or None,
        has_power=has_power,
        power_peak_w=power_peak_w,
        power_nominal_w=power_nominal_w,
        packing=packing_input,
        accessories=accessory_items,
    )
    return data, total_quantity


def _find_by_name(objects, name: str):
    """Найти объект по названию без учёта регистра (Python — SQL lower() не
    гарантирует корректную работу с кириллицей на SQLite)."""
    target = name.strip().lower()
    for obj in objects:
        if obj.name.strip().lower() == target:
            return obj
    return None


def _read_accessories(ws, db: Session, errors: list[str]) -> list[AccessoryQty]:
    header_row = None
    for r in range(1, ws.max_row + 1):
        if (
            _cell_str(ws, r, 1) == ACCESSORY_HEADER[0]
            and _cell_str(ws, r, 2) == ACCESSORY_HEADER[1]
        ):
            header_row = r
            break
    if header_row is None:
        return []

    accessory_categories = acc_service.list_accessory_categories(db)
    result: list[AccessoryQty] = []
    r = header_row + 1
    while r <= ws.max_row:
        cat_name = _cell_str(ws, r, 1)
        acc_name = _cell_str(ws, r, 2)
        qty_raw = _cell_str(ws, r, 3)
        if not cat_name and not acc_name and not qty_raw:
            break
        if not cat_name or not acc_name:
            errors.append(f"Комплектация, строка {r}: не заполнены категория и/или аксессуар.")
            r += 1
            continue
        category = _find_by_name(accessory_categories, cat_name)
        if category is None:
            errors.append(
                f"Комплектация, строка {r}: категория аксессуара «{cat_name}» не найдена."
            )
            r += 1
            continue
        accessory = _find_by_name(category.accessories, acc_name)
        if accessory is None:
            errors.append(
                f"Комплектация, строка {r}: аксессуар «{acc_name}» не найден "
                f"в категории «{cat_name}»."
            )
            r += 1
            continue
        quantity = _parse_int_field(
            qty_raw, "Комплектация: Количество", errors, minimum=1, default=1
        )
        result.append(AccessoryQty(accessory_id=accessory.id, quantity=quantity))
        r += 1
    return result


def _read_items(
    ws, db: Session, model: EquipmentModel, errors: list[str]
) -> list[_ItemPlan]:
    items_by_id = {it.id: it for it in item_service.list_items(db, model.id)}

    existing_barcode_owner: dict[str, int] = {}
    for iid, barcode in db.execute(
        select(EquipmentItem.id, EquipmentItem.barcode).where(EquipmentItem.barcode.isnot(None))
    ):
        existing_barcode_owner[item_service.normalize_barcode(barcode)] = iid

    seen_ids: set[int] = set()
    claimed_in_file: dict[str, int] = {}
    plans: list[_ItemPlan] = []

    row = 2
    while row <= ws.max_row:
        # Колонка 1 («Модель») — информационная, не участвует в сопоставлении.
        id_raw = _cell_str(ws, row, 2)
        barcode_raw = _cell_str(ws, row, 3)
        serial_raw = _cell_str(ws, row, 4)
        inv_raw = _cell_str(ws, row, 5)
        status_raw = _cell_str(ws, row, 6)
        usable_raw = _cell_str(ws, row, 7)
        comment_raw = _cell_str(ws, row, 8)
        if not any([id_raw, barcode_raw, serial_raw, inv_raw, status_raw, usable_raw, comment_raw]):
            row += 1
            continue

        existing_item: EquipmentItem | None = None
        if id_raw:
            try:
                iid = int(float(id_raw))
            except ValueError:
                errors.append(f"Единицы, строка {row}: некорректный ID «{id_raw}».")
                row += 1
                continue
            if iid in seen_ids:
                errors.append(f"Единицы, строка {row}: ID {iid} повторяется в файле.")
                row += 1
                continue
            seen_ids.add(iid)
            existing_item = items_by_id.get(iid)
            if existing_item is None:
                errors.append(
                    f"Единицы, строка {row}: единица с ID {iid} не найдена у этой модели."
                )
                row += 1
                continue

        barcode = barcode_raw or None
        if barcode:
            norm = item_service.normalize_barcode(barcode)
            own_id = existing_item.id if existing_item else None
            owner = existing_barcode_owner.get(norm)
            if owner is not None and owner != own_id:
                errors.append(
                    f"Единицы, строка {row}: штрих-код «{barcode}» уже используется другой "
                    f"единицей (ID {owner})."
                )
            claim_row = claimed_in_file.get(norm)
            if claim_row is not None:
                errors.append(
                    f"Единицы, строка {row}: штрих-код «{barcode}» повторяется в файле "
                    f"(см. строку {claim_row})."
                )
            else:
                claimed_in_file[norm] = row

        status_key = (status_raw or ItemStatus.ACTIVE.label).strip().lower()
        status = STATUS_BY_LABEL.get(status_key)
        if status is None:
            errors.append(f"Единицы, строка {row}: неизвестный статус «{status_raw}».")
            status = ItemStatus.ACTIVE

        plans.append(
            _ItemPlan(
                row=row,
                existing=existing_item,
                barcode=barcode,
                serial_number=serial_raw or None,
                inventory_number=inv_raw or None,
                comment=comment_raw or None,
                status=status,
                usable_despite_defect=_parse_bool(usable_raw),
            )
        )
        row += 1

    return plans


# --- Импорт: применение -----------------------------------------------------


def import_workbook(
    db: Session, model: EquipmentModel, file_bytes: bytes, user_id: int | None
) -> ImportResult:
    wb = _load_workbook(file_bytes)

    if CARD_SHEET not in wb.sheetnames or ITEMS_SHEET not in wb.sheetnames:
        raise ImportValidationError(
            [f"Файл должен содержать листы «{CARD_SHEET}» и «{ITEMS_SHEET}»."]
        )

    errors: list[str] = []
    data, total_quantity = _read_card_fields(wb[CARD_SHEET], db, model, errors)
    item_plans = _read_items(wb[ITEMS_SHEET], db, model, errors)

    if errors or data is None:
        raise ImportValidationError(errors)

    eq_service.update_model(db, model, data)

    items_created = items_updated = 0
    for plan in item_plans:
        if plan.existing is not None:
            it = plan.existing
            changed = (
                it.barcode != plan.barcode
                or it.serial_number != plan.serial_number
                or it.inventory_number != plan.inventory_number
                or it.comment != plan.comment
            )
            it.barcode = plan.barcode
            it.serial_number = plan.serial_number
            it.inventory_number = plan.inventory_number
            it.comment = plan.comment
            if plan.status != it.status:
                it.status_history.append(
                    EquipmentStatusHistory(
                        changed_at=utcnow(),
                        user_id=user_id,
                        old_status=it.status,
                        new_status=plan.status,
                        comment="Импорт из Excel",
                    )
                )
                it.status = plan.status
                it.usable_despite_defect = (
                    plan.usable_despite_defect if plan.status == ItemStatus.DEFECT else False
                )
                changed = True
            elif (
                plan.status == ItemStatus.DEFECT
                and it.usable_despite_defect != plan.usable_despite_defect
            ):
                it.usable_despite_defect = plan.usable_despite_defect
                changed = True
            if changed:
                items_updated += 1
        else:
            new_item = EquipmentItem(
                model_id=model.id,
                barcode=plan.barcode,
                status=plan.status,
                serial_number=plan.serial_number,
                inventory_number=plan.inventory_number,
                comment=plan.comment,
                usable_despite_defect=(
                    plan.usable_despite_defect if plan.status == ItemStatus.DEFECT else False
                ),
            )
            new_item.status_history.append(
                EquipmentStatusHistory(
                    changed_at=utcnow(),
                    user_id=user_id,
                    old_status=None,
                    new_status=plan.status,
                    comment="Импорт из Excel",
                )
            )
            db.add(new_item)
            items_created += 1

    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise ImportValidationError(
            ["Не удалось сохранить единицы оборудования — конфликт штрих-кода. Повторите импорт."]
        ) from exc
    db.refresh(model)

    quantity_adjusted = False
    if model.accounting_type == AccountingType.QUANTITY:
        current = eq_service.active_count(db, model.id)
        if total_quantity != current:
            try:
                eq_service.adjust_quantity(
                    db, model, total_quantity, user_id, comment="Импорт из Excel"
                )
                quantity_adjusted = True
            except InventoryError:
                pass

    return ImportResult(
        items_created=items_created,
        items_updated=items_updated,
        quantity_adjusted=quantity_adjusted,
    )
